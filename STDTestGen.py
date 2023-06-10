import sys
import os
import shutil
import datetime
import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QPushButton, QVBoxLayout, QHBoxLayout, QComboBox, QSpinBox, QFileDialog, QMessageBox
from PyQt5.QtCore import Qt, QRegExp, QSize
from PyQt5.QtGui import QRegExpValidator, QPixmap, QIcon, QFont

class TestRequestApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('STD Test Generator')
        self.setFixedSize(500, 650)
        font = QFont("Segoe UI", 10)
        QApplication.instance().setFont(font)         
        layout = QVBoxLayout()
        layout.addSpacing(10)

        # Item 1
        self.sol_input = QLineEdit()
        self.sol_input.setMaxLength(8)
        regex = QRegExp(r'\d{3}/\d{4}')
        validator = QRegExpValidator(regex)
        self.sol_input.setValidator(validator)
        self.sol_input.textChanged.connect(self.insert_slash)
        search_button = QPushButton('Buscar')
        search_button.clicked.connect(self.search)

        hbox1 = QHBoxLayout()
        hbox1.addWidget(QLabel('Solicitação de Teste:'))
        hbox1.addWidget(self.sol_input)
        hbox1.addWidget(search_button)
        
        # adiciona um espaçador vertical de 10 pixels entre as linhas horizontais
        layout.addLayout(hbox1)
        layout.addSpacing(10)

        # Item 1.1 - 1.4
        self.linha_label = QLabel('Linha:')
        self.linha_input = QLineEdit()
        self.linha_input.setFixedWidth(420)
        self.linha_input.setReadOnly(True)
        self.linha_input.setStyleSheet('background-color: lightgray')
        hbox2 = QHBoxLayout()
        hbox2.addWidget(self.linha_label)
        hbox2.addWidget(self.linha_input)
        layout.addLayout(hbox2)

        self.familia_label = QLabel('Familia:')
        self.familia_input = QLineEdit()
        self.familia_input.setFixedWidth(420)
        self.familia_input.setReadOnly(True)
        self.familia_input.setStyleSheet('background-color: lightgray')
        hbox3 = QHBoxLayout()
        hbox3.addWidget(self.familia_label)
        hbox3.addWidget(self.familia_input)
        layout.addLayout(hbox3)

        self.projeto_label = QLabel('Projeto:')
        self.projeto_input = QLineEdit()
        self.projeto_input.setFixedWidth(420)
        self.projeto_input.setReadOnly(True)
        self.projeto_input.setStyleSheet('background-color: lightgray')
        hbox4 = QHBoxLayout()
        hbox4.addWidget(self.projeto_label)
        hbox4.addWidget(self.projeto_input)
        layout.addLayout(hbox4)

        self.modelo_label = QLabel('Modelo:')
        self.modelo_input = QLineEdit()
        self.modelo_input.setFixedWidth(420)
        self.modelo_input.setReadOnly(True)
        self.modelo_input.setStyleSheet('background-color: lightgray')
        hbox5 = QHBoxLayout()
        hbox5.addWidget(self.modelo_label)
        hbox5.addWidget(self.modelo_input)
        layout.addLayout(hbox5)
        layout.addSpacing(20)

        # Item 2
        self.test_selection_label = QLabel()
        self.test_selection = QComboBox()
        self.test_selection.addItem('')
        self.populate_test_selection()
        self.test_selection.currentIndexChanged.connect(self.test_selected)
        hbox6 = QHBoxLayout()
        hbox6.addWidget(QLabel('Selecionar ensaio:'))
        hbox6.addWidget(self.test_selection)
        layout.addLayout(hbox6)

        # Item 3
        self.frf_label = QLabel()
        self.frf_selection = QComboBox()
        self.frf_selection.addItem('')
        self.frf_selection.setEnabled(False)
        hbox7 = QHBoxLayout()
        hbox7.addWidget(QLabel('FRF:'))
        hbox7.addWidget(self.frf_selection)
        layout.addLayout(hbox7)

        # Item 4
        self.model_selection_label = QLabel()
        self.model_selection = QComboBox()
        hbox8 = QHBoxLayout()
        hbox8.addWidget(QLabel('Modelo:'))
        hbox8.addWidget(self.model_selection)
        layout.addLayout(hbox8)

        # Item 5
        self.num_samples_label = QLabel()
        self.num_samples = QSpinBox()
        self.num_samples.setRange(0, 5)
        hbox9 = QHBoxLayout()
        hbox9.addWidget(QLabel('Número de amostras:'))
        hbox9.addWidget(self.num_samples)
        layout.addLayout(hbox9)

        # Item 6
        # define the method convert_to_upper
        def convert_to_upper(text, index):
            # convert any lowercase letter to uppercase
            self.ns_inputs[index].setText(text.upper())

        self.ns_inputs = []
        layout.addSpacing(10)
        for i in range(5):
            ns_input = QLineEdit()
            ns_input.setFixedWidth(236)
            ns_input.setPlaceholderText(f" N/S {str(i + 1)}")
            ns_input.hide()
            # set the regular expression validator
            validator = QRegExpValidator(QRegExp("[A-Z0-9]*"), ns_input)
            ns_input.setValidator(validator)
            # connect the textChanged event to the uppercase conversion function
            ns_input.textChanged.connect(lambda text, i=i: convert_to_upper(text, i))
            self.ns_inputs.append(ns_input)
            hbox10 = QHBoxLayout()
            hbox10.addWidget(QLabel(' '))
            hbox10.addWidget(self.ns_inputs[i])
            layout.addLayout(hbox10)
            self.setLayout(layout)
            self.num_samples.valueChanged.connect(self.update_ns_inputs)
            
        # layout.addSpacing(500)
        # Item 7
        layout.addSpacing(200)
        create_button = QPushButton('CRIAR DOCUMENTOS DE TESTE  ')
        create_button.clicked.connect(self.create_folders_and_documents)
        create_button.setStyleSheet('font-weight: bold; height: 35px; width: 240px')
        # Set the icon for the button
        #ABAIXO, DIRETÓRIO PC MATEUS
        icon_path = 'C:\\Users\\Mateus\\Documents\\Projetos\\STD Test Gen\\Engenharia de Produto - LAB\\01-Templates\\logo\\plus_doc.png'
        #ABAIXO, DIRETÓRIO PC LAB
        # icon_path = 'C:\\Users\\ulabeng\\Midea Carrier\\Engenharia de Produto - LAB\\01-Templates\\logo\\plus_doc.png'
        create_button.setIcon(QIcon(icon_path))
        create_button.setIconSize(QSize(32, 32))  
        hbox11 = QHBoxLayout()
        hbox11.addStretch() 
        hbox11.addWidget(create_button)
        hbox11.addStretch() 
        layout.addLayout(hbox11)

        # Logo
        #ABAIXO, DIRETÓRIO PC MATEUS
        logo = QPixmap('C:\\Users\\Mateus\\Documents\\Projetos\\STD Test Gen\\Engenharia de Produto - LAB\\01-Templates\\logo\\logo.png')
        #ABAIXO, DIRETÓRIO PC LAB
        # logo = QPixmap('C:\\Users\\ulabeng\\Midea Carrier\\Engenharia de Produto - LAB\\01-Templates\\logo\\logo.png')
        logo = logo.scaled(175, 85, Qt.KeepAspectRatio)
        logo_label = QLabel()
        logo_label.setPixmap(logo)
        logo_label.setAlignment(Qt.AlignCenter)
        hbox12 = QHBoxLayout()
        hbox12.addStretch()  
        hbox12.addWidget(logo_label)
        hbox12.addStretch()  
        layout.addLayout(hbox12)

        developer_label = QLabel('developer: mcaregnatto - v1.2')
        developer_label.setAlignment(Qt.AlignCenter)
        developer_label.setStyleSheet('color: gray; font-size: 9px')
        hbox13 = QHBoxLayout()
        hbox13.addStretch()  
        hbox13.addWidget(developer_label)
        hbox13.addStretch()  
        layout.addLayout(hbox13)

    def insert_slash(self, text):
        if len(text) == 3 and not text.endswith('/'):
            self.sol_input.setText(text + '/')

    def search(self):
        sol = self.sol_input.text()
        if not sol:
            QMessageBox.warning(self, 'Aviso', 'Solicitação não existente')
            return

        #ABAIXO, DIRETÓRIO PC MATEUS
        wb = openpyxl.load_workbook('C:\\Users\\Mateus\\Documents\\Projetos\\STD Test Gen\\Engenharia de Produto - LAB\\00-Solicitações de Teste\\SOLICITAÇÕES DE TESTE.xlsx')
        #ABAIXO, DIRETÓRIO PC LAB
        # wb = openpyxl.load_workbook('C:\\Users\\ulabeng\\Midea Carrier\\Engenharia de Produto - LAB\\00-Solicitações de Teste\\SOLICITAÇÕES DE TESTE.xlsx')
        ws = wb.active

        found = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == sol:
                found = True
                self.linha_input.setText(row[3].value)
                self.familia_input.setText(row[4].value)
                self.projeto_input.setText(row[5].value)
                self.modelo_input.setText(f'{row[8].value} x {row[7].value}')
                self.model_selection.addItem(row[8].value)
                self.model_selection.addItem(row[7].value)
                break

        if not found:
            QMessageBox.warning(self, 'Aviso', 'Solicitação não existente')

    def populate_test_selection(self):
        #ABAIXO, DIRETÓRIO PC MATEUS
        path = os.path.join('C:\\Users\\Mateus\\Documents\\Projetos\\STD Test Gen\\Engenharia de Produto - LAB\\01-Templates\\main')
        #ABAIXO, DIRETÓRIO PC LAB
        # path = os.path.join('C:\\Users\\ulabeng\\Midea Carrier\\Engenharia de Produto - LAB\\01-Templates\\main')
        for file in os.listdir(path):
            if file.endswith('.xlsx'):
                self.test_selection.addItem(file[:-5])

    def test_selected(self, index):
        selected_test = self.test_selection.itemText(index)
        if ('Variable' in selected_test):
            self.frf_selection.setEnabled(True)
            self.populate_frf_selection()
        else:
            self.frf_selection.setEnabled(False)
            self.frf_selection.clear()
            self.frf_selection.addItem('')

    def populate_frf_selection(self):
        #ABAIXO, DIRETÓRIO PC MATEUS
        path = os.path.join('C:\\Users\\Mateus\\Documents\\Projetos\\STD Test Gen\\Engenharia de Produto - LAB\\01-Templates\\frf')
        #ABAIXO, DIRETÓRIO PC LAB
        # path = os.path.join('C:\\Users\\ulabeng\\Midea Carrier\\Engenharia de Produto - LAB\\01-Templates\\frf')
        self.frf_selection.clear()
        self.frf_selection.addItem('')
        for file in os.listdir(path):
            if file.endswith('.xlsx'):
                self.frf_selection.addItem(file[:-5])

    def update_ns_inputs(self, value):
        for i in range(5):
            if i < value:
               # self.ns_labels[i].show()
                self.ns_inputs[i].show()
            else:
                #self.ns_labels[i].hide()
                self.ns_inputs[i].hide()

    def create_folders_and_documents(self):
        sol_traco = self.sol_input.text().replace('/', '-')
        kind_test = self.test_selection.currentText()[5:].split('(')[0].strip()
        linha = self.linha_input.text()
        familia = self.familia_input.text()
        projeto = self.projeto_input.text()
        modelo = self.model_selection.currentText()
        modelo_evcd = self.modelo_input.text()
        num_samples = self.num_samples.value()
        entered_serial_numbers = []

        if not kind_test or not linha or not familia or not projeto or not modelo:
            QMessageBox.warning(self, 'Aviso', 'Preencha todos os campos obrigatórios')
            return
        
        if num_samples == 0:
            QMessageBox.warning(self, 'Aviso', 'O número de amostras deve ser maior que zero')
            return
        
        #ABAIXO, DIRETÓRIO PC MATEUS
        base_path = 'C:\\Users\\Mateus\\Documents\\Projetos\\STD Test Gen\\Engenharia de Produto - LAB'
        #ABAIXO, DIRETÓRIO PC LAB
        # base_path = 'C:\\Users\\ulabeng\\Midea Carrier\\Engenharia de Produto - LAB'
        
        # Excption if "sonora"
        if "sonora" in kind_test:
            path = os.path.join(base_path, "Ruído", linha, familia, projeto, modelo_evcd)
        else:
            path = os.path.join(base_path, kind_test, linha, familia, projeto, modelo_evcd)
        
        os.makedirs(path, exist_ok=True)

        # Adicionando a pasta vazia "sol_traco" + "-" + "Dados de teste"
        test_data_folder = os.path.join(path, f"{sol_traco} - Dados de teste")
        os.makedirs(test_data_folder, exist_ok=True)
        today = datetime.datetime.now().strftime('%Y%m%d')

        for i in range(num_samples):
            ns = self.ns_inputs[i].text()
            if not ns:
                QMessageBox.warning(self, 'Aviso', f'Preencha o campo N/S {i + 1}')
                return
                # Verificando se o número de série já foi inserido
            if ns in entered_serial_numbers:
                QMessageBox.warning(self, 'Aviso', f'Número de série repetido.')
                return
            else:
                entered_serial_numbers.append(ns)
            
            #ABAIXO, DIRETÓRIO PC MATEUS
            src_file = os.path.join('C:\\Users\\Mateus\\Documents\\Projetos\\STD Test Gen\\Engenharia de Produto - LAB\\01-Templates\\main', f'{self.test_selection.currentText()}.xlsx')
            #ABAIXO, DIRETÓRIO PC LAB
            # src_file = os.path.join('C:\\Users\\ulabeng\\Midea Carrier\\Engenharia de Produto - LAB\\01-Templates\\main', f'{self.test_selection.currentText()}.xlsx')
            dst_file = os.path.join(path, f'{today}_{sol_traco}_{modelo}_{ns}_{kind_test}.xlsx')
            shutil.copy(src_file, dst_file)

            if self.frf_selection.isEnabled():
                #ABAIXO, DIRETÓRIO PC MATEUS
                src_file = os.path.join('C:\\Users\\Mateus\\Documents\\Projetos\\STD Test Gen\\Engenharia de Produto - LAB\\01-Templates\\frf', f'{self.frf_selection.currentText()}.xlsx')
                #ABAIXO, DIRETÓRIO PC LAB
                # src_file = os.path.join('C:\\Users\\ulabeng\\Midea Carrier\\Engenharia de Produto - LAB\\01-Templates\\frf', f'{self.frf_selection.currentText()}.xlsx')
                dst_file_frf = os.path.join(path, f'{today}_{sol_traco}_{modelo}_{ns}_{kind_test}_FRF.xlsx') 
                shutil.copy(src_file, dst_file_frf)  # Copie o arquivo FRF para o novo arquivo de destino

        # Create shortcut
        shortcut_path = os.path.join(os.path.expanduser('~'), 'Desktop', f'{kind_test}_{sol_traco}_{linha}_{familia}_{projeto}_{modelo}.lnk')
        target_path = os.path.abspath(path)
        self.create_shortcut(shortcut_path, target_path)
        QMessageBox.information(self, 'Aviso', 'Concluído com sucesso!')

    def create_shortcut(self, shortcut_path, target_path):
        import pythoncom
        from win32com.shell import shell
        shortcut = pythoncom.CoCreateInstance(
            shell.CLSID_ShellLink,
            None,
         pythoncom.CLSCTX_INPROC_SERVER,
            shell.IID_IShellLink
        )

        shortcut.SetPath(target_path)
        shortcut.SetDescription(f'Shortcut to {target_path}')
        shortcut.SetIconLocation(target_path, 0)
        persist_file = shortcut.QueryInterface(pythoncom.IID_IPersistFile)
        persist_file.Save(shortcut_path, 0)
    
if __name__ == '__main__':
    app = QApplication(sys.argv)
    # Define o nome do aplicativo
    app.setApplicationDisplayName('STD Test Generator')
    ex = TestRequestApp()
    ex.show()
    ex.setWindowIcon(QIcon('C:\\Users\\Mateus\\Documents\\Projetos\\STD Test Gen\\Engenharia de Produto - LAB\\01-Templates\\ico16.ico'))
    sys.exit(app.exec_())